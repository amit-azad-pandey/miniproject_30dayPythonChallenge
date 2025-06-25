from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from pathlib import Path

app = FastAPI()

# Mount static files and templates
templates = Jinja2Templates(directory="templates")

EXCEL_FILE = "expenses.xlsx"

# Create Excel file if it doesn't exist
if not Path(EXCEL_FILE).exists():
    wb = Workbook()
    ws = wb.active
    ws.append(["Salary", "Category", "Amount (Rs)", "Description", "Paid Via"])
    wb.save(EXCEL_FILE)

@app.get("/", response_class=HTMLResponse)
def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/add_expense")
def add_expense(salary: float = Form(...), category: str = Form(...),
                amount: float = Form(...), description: str = Form(...),
                paid_via: str = Form(...)):
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([salary, category, amount, description, paid_via])
    wb.save(EXCEL_FILE)

    return RedirectResponse("/", status_code=303)
